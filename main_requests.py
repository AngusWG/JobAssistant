#!/usr/bin/python3
# encoding: utf-8 
# @Time    : 2018/9/20 17:09 
# @author  : zza
# @Email   : 740713651@qq.com
import random
import re
import time
from multiprocessing import Pool
from urllib import parse
from openpyxl import load_workbook, Workbook
from selenium import webdriver
import os

import requests
from lxml import etree
import config
from 代理ip池 import 代理ip池

header = {"User-Agent": "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1 Trident/5.0;",
          "Cookie": "index_location_city=" + parse.quote(config.city) + ";"
          }

key_list = config.关键字.strip().split()
key_matching_num = int(config.关键词匹配量)

ip_pool = 代理ip池()
ip_pool.main()

time.sleep(10)


class LaGou():
    # user_cookies = "C:\\Users\\74071\\AppData\\Local\\Google\\Chrome\\User Data"
    # option = webdriver.ChromeOptions()
    # option.add_argument("--user-data-dir={}".format(user_cookies))  # 设置成用户自己的数据目录
    # browser = webdriver.Chrome(chrome_options=option)

    main_url = "https://www.lagou.com/zhaopin/"

    def __init__(self):
        # self.browser.get(self.main_url)
        self.crawl_list = config.爬取列表.strip().split()
        self.max_item_num = int(config.查找最高量)
        self.result = []
        self.page_result = []
        pass

    def server(self):
        self.list_search()
        self.list_page_confirm()
        self.save_2_doc()
        print("over")

    def save_2_doc(self):
        if os.path.exists('doc/LaGou.xls'):
            os.remove("doc/LaGou.xls")
        wb = Workbook()
        ws = wb.create_sheet()
        for i in range(len(self.page_result)):
            if self.page_result[i] is None:
                continue
            ws.append(self.page_result[i])
        wb.save('doc/LaGou.xlsx')
        print("{}条 保存到doc文件夹下 请查收 ".format(len(self.page_result)))

    @staticmethod
    def list_search_item(main_url, item, proxies):
        url_list = list()
        url = main_url + item + "/{}/"
        file_name = os.path.join("data", item + ".txt")
        # 检查历史防中断
        try:
            with open(file_name, "r", encoding="utf8") as f:
                url_list.extend(f.read().splitlines())
                first_page = int(len(url_list) / 15) + 1
        except:
            first_page = 1

        # 爬取
        try:
            for page_num in range(first_page, 31):
                req1 = requests.get(url.format(page_num), headers=header, proxies=proxies)
                if req1.status_code != 200:
                    print("{}：{}页面没有数据 over".format(item, page_num))
                    break
                get_item = etree.HTML(req1.text).xpath('//*[@id="s_position_list"]/ul//div[1]/div[1]/a[1]/@href')
                if len(get_item) == 0:
                    raise Exception("被墙了")
                url_list.extend(get_item)
                print("{}:第{}页 已获得{}条待爬取页面".format(item, page_num, len(url_list)))
                break  # test
        except Exception as err:
            print(err)
        finally:
            # 保存历史
            with open(file_name, "w", encoding="utf8") as f:
                f.writelines("\n".join(url_list))
        return url_list

    def list_search(self):
        pool = Pool(6)
        for item in self.crawl_list:
            print(random.choice(list(ip_pool.ip代理池)))
            proxies = {
                "https": random.choice(list(ip_pool.ip代理池)),
            }
            self.result.append(pool.apply_async(self.list_search_item, args=(self.main_url, item, proxies)))
        pool.close()
        pool.join()
        # 去重
        result = []
        [result.extend(i.get()) for i in self.result]
        self.result = list(set(result))
        print(len(self.result))
        print("list_search over", "*" * 60)

    def list_page_confirm(self):
        # 每页扫描
        pool = Pool(6)
        for item_url in self.result:
            self.page_result.append(pool.apply_async(self.page_confirm, args=(item_url,)))
        pool.close()
        pool.join()
        # 存储结果
        self.page_result = [i.get() for i in self.page_result]
        with open("./data/data.txt", "w", encoding="utf8") as f:
            for item in self.page_result:
                if item is None:
                    continue
                f.writelines(str(item) + "\n")
        print(len(self.page_result))
        print("list_page_confirm over", "*" * 60)

    @staticmethod
    def page_confirm(item_url):
        result = ["投", item_url]
        # 获取数据
        req1 = requests.get(item_url, headers=header)
        if req1.status_code != 200:
            print("{}页面没有数据 over".format(item_url))
            return
        selector = etree.HTML(req1.text)
        # 匹配关键字
        text = selector.xpath('//*[@id="job_detail"]//text()')
        res = re.findall("|".join(key_list), "".join(text), flags=re.IGNORECASE)
        res = list(set([i.lower() for i in res]))
        if len(res) < key_matching_num:
            return
        # 获取信息
        try:
            # 要求
            text1 = " ".join("".join(selector.xpath('//*[@class="job_bt"]//text()')).split())
            result.append(text1)
            # 薪资相关
            text2 = " ".join("".join(selector.xpath('//*[@class="job_request"]//text()')).split())
            result.append(text2)
            # 公司名字
            text3 = " ".join("".join(selector.xpath('//*[@class="company"]//text()')).split())
            result.append(text3)
            print("page over", item_url, res)
            return result
        except Exception as err:
            raise err


def main():
    while len(list(ip_pool.ip代理池)) < 5:
        time.sleep(5)
    # LaGou().make_doc()
    c = LaGou()
    c.server()
    pass


if __name__ == '__main__':
    main()
